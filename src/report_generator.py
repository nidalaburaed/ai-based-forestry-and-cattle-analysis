"""
Forest Plan Report Generator
=============================

Generates professional forestry reports in multiple formats:
- PDF: Formal forestry plan document
- Excel: Detailed data tables
- HTML: Interactive web report
- CSV: Data export for GIS/analysis tools

Author: Adrian
Date: 2024-12-15
"""

import json
from pathlib import Path
from datetime import datetime
from typing import Dict, Any, Optional, Union
import logging

logger = logging.getLogger(__name__)


class ReportGenerator:
    """Generate professional forestry reports in multiple formats."""
    
    def __init__(self, output_dir: str = "output"):
        """Initialize report generator.
        
        Args:
            output_dir: Directory to save reports
        """
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
    def generate_all_reports(self, forest_plan, aggregated_stats: Dict[str, Any],
                            video_path: Union[str, int]) -> Dict[str, Path]:
        """Generate all report formats.

        Args:
            forest_plan: ForestPlan object
            aggregated_stats: Aggregated statistics dictionary
            video_path: Path to analyzed video or camera index

        Returns:
            Dictionary mapping format to output file path
        """
        # Convert video_path to string for display
        video_path_str = str(video_path)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        base_name = f"forest_report_{timestamp}"

        outputs = {}

        # Generate HTML report (always available)
        html_path = self.generate_html_report(forest_plan, aggregated_stats,
                                              video_path_str, base_name)
        outputs['html'] = html_path
        
        # Generate CSV exports
        csv_paths = self.generate_csv_exports(forest_plan, aggregated_stats, base_name)
        outputs.update(csv_paths)
        
        # Try to generate PDF (requires reportlab)
        try:
            pdf_path = self.generate_pdf_report(forest_plan, aggregated_stats,
                                                video_path_str, base_name)
            outputs['pdf'] = pdf_path
        except ImportError:
            logger.warning("PDF generation requires 'reportlab' package. Skipping PDF.")
        except Exception as e:
            logger.error(f"Error generating PDF: {e}")

        # Try to generate Excel (requires openpyxl)
        try:
            excel_path = self.generate_excel_report(forest_plan, aggregated_stats,
                                                    video_path_str, base_name)
            outputs['excel'] = excel_path
        except ImportError:
            logger.warning("Excel generation requires 'openpyxl' package. Skipping Excel.")
        except Exception as e:
            logger.error(f"Error generating Excel: {e}")
            
        return outputs
    
    def generate_html_report(self, forest_plan, aggregated_stats: Dict[str, Any],
                            video_path: str, base_name: str) -> Path:
        """Generate HTML report with embedded CSS and JavaScript.
        
        Args:
            forest_plan: ForestPlan object
            aggregated_stats: Aggregated statistics dictionary
            video_path: Path to analyzed video
            base_name: Base filename for output
            
        Returns:
            Path to generated HTML file
        """
        output_path = self.output_dir / f"{base_name}.html"
        
        # Build HTML content
        html_content = self._build_html_content(forest_plan, aggregated_stats, video_path)
        
        # Write to file
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(html_content)
            
        logger.info(f"HTML report generated: {output_path}")
        return output_path
    
    def generate_csv_exports(self, forest_plan, aggregated_stats: Dict[str, Any],
                            base_name: str) -> Dict[str, Path]:
        """Generate CSV exports for sites, stands, and activities.
        
        Args:
            forest_plan: ForestPlan object
            aggregated_stats: Aggregated statistics dictionary
            base_name: Base filename for output
            
        Returns:
            Dictionary mapping CSV type to file path
        """
        import csv
        
        outputs = {}
        
        # Export sites
        if forest_plan.sites:
            sites_path = self.output_dir / f"{base_name}_sites.csv"
            with open(sites_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=['name', 'type', 'nutrient_level', 
                                                       'timber_potential', 'notes'])
                writer.writeheader()
                for site in forest_plan.sites:
                    writer.writerow({
                        'name': site.name,
                        'type': site.type,
                        'nutrient_level': site.nutrient_level,
                        'timber_potential': site.timber_potential,
                        'notes': site.notes or ''
                    })
            outputs['sites_csv'] = sites_path
            logger.info(f"Sites CSV generated: {sites_path}")

        # Export stands
        if forest_plan.stands:
            stands_path = self.output_dir / f"{base_name}_stands.csv"
            with open(stands_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=['species', 'age', 'size_ha',
                                                       'development_class', 'notes'])
                writer.writeheader()
                for stand in forest_plan.stands:
                    writer.writerow({
                        'species': ', '.join(stand.species),
                        'age': stand.age or '',
                        'size_ha': stand.size_ha or '',
                        'development_class': stand.development_class or '',
                        'notes': stand.notes or ''
                    })
            outputs['stands_csv'] = stands_path
            logger.info(f"Stands CSV generated: {stands_path}")

        # Export management activities
        if forest_plan.management_activities:
            activities_path = self.output_dir / f"{base_name}_activities.csv"
            with open(activities_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=['description', 'start_date',
                                                       'end_date', 'responsible_party', 'notes'])
                writer.writeheader()
                for activity in forest_plan.management_activities:
                    writer.writerow({
                        'description': activity.description,
                        'start_date': str(activity.start_date) if activity.start_date else '',
                        'end_date': str(activity.end_date) if activity.end_date else '',
                        'responsible_party': activity.responsible_party or '',
                        'notes': activity.notes or ''
                    })
            outputs['activities_csv'] = activities_path
            logger.info(f"Activities CSV generated: {activities_path}")

        # Export aggregated statistics
        stats_path = self.output_dir / f"{base_name}_statistics.csv"
        with open(stats_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(['Metric', 'Value'])
            for key, value in aggregated_stats.items():
                if key != 'species_distribution':  # Handle separately
                    writer.writerow([key, value])

            # Add species distribution
            if 'species_distribution' in aggregated_stats:
                writer.writerow(['', ''])  # Empty row
                writer.writerow(['Species Distribution', ''])
                for species, count in aggregated_stats['species_distribution'].items():
                    writer.writerow([species, count])
        outputs['statistics_csv'] = stats_path
        logger.info(f"Statistics CSV generated: {stats_path}")

        return outputs

    def generate_pdf_report(self, forest_plan, aggregated_stats: Dict[str, Any],
                           video_path: str, base_name: str) -> Path:
        """Generate PDF report (requires reportlab).

        Args:
            forest_plan: ForestPlan object
            aggregated_stats: Aggregated statistics dictionary
            video_path: Path to analyzed video
            base_name: Base filename for output

        Returns:
            Path to generated PDF file
        """
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER, TA_LEFT

        output_path = self.output_dir / f"{base_name}.pdf"

        # Create PDF document
        doc = SimpleDocTemplate(str(output_path), pagesize=A4,
                               rightMargin=2*cm, leftMargin=2*cm,
                               topMargin=2*cm, bottomMargin=2*cm)

        # Container for PDF elements
        elements = []
        styles = getSampleStyleSheet()

        # Custom styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#2d3748'),
            spaceAfter=30,
            alignment=TA_CENTER
        )

        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor('#2d3748'),
            spaceAfter=12,
            spaceBefore=12
        )

        # Title
        elements.append(Paragraph(f"Forest Analysis Report", title_style))
        elements.append(Paragraph(f"{forest_plan.name}", styles['Heading2']))
        elements.append(Spacer(1, 0.5*cm))

        # Analysis metadata
        metadata_data = [
            ['Analysis Date:', aggregated_stats.get('analysis_date', 'N/A')],
            ['Video Source:', Path(video_path).name],
            ['Frames Analyzed:', str(aggregated_stats.get('total_frames_analyzed', 0))],
            ['Report Generated:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
        ]
        metadata_table = Table(metadata_data, colWidths=[5*cm, 10*cm])
        metadata_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f7fafc')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#2d3748')),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
        ]))
        elements.append(metadata_table)
        elements.append(Spacer(1, 1*cm))

        # Executive Summary - Aggregated Statistics
        elements.append(Paragraph("Executive Summary", heading_style))

        stats_data = [
            ['Metric', 'Value'],
            ['Total Trees Detected', str(aggregated_stats.get('total_trees_detected', 0))],
            ['Avg Trees per Frame', f"{aggregated_stats.get('avg_trees_per_frame', 0):.2f}"],
            ['Canopy Coverage', f"{aggregated_stats.get('avg_canopy_coverage_percent', 0):.2f}%"],
            ['Vegetation Density', f"{aggregated_stats.get('avg_vegetation_density_percent', 0):.2f}%"],
        ]

        if aggregated_stats.get('avg_tree_height_m'):
            stats_data.append(['Avg Tree Height', f"{aggregated_stats['avg_tree_height_m']:.2f} m"])
        if aggregated_stats.get('avg_dbh_cm'):
            stats_data.append(['Avg DBH', f"{aggregated_stats['avg_dbh_cm']:.2f} cm"])

        stats_table = Table(stats_data, colWidths=[8*cm, 7*cm])
        stats_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(stats_table)
        elements.append(Spacer(1, 1*cm))

        # Sites Section
        if forest_plan.sites:
            elements.append(Paragraph("Sites", heading_style))
            for i, site in enumerate(forest_plan.sites, 1):
                site_text = f"<b>Site {i}:</b> {site.name or 'Unnamed'}<br/>"
                site_text += f"Type: {site.type}, Nutrient Level: {site.nutrient_level}, "
                site_text += f"Timber Potential: {site.timber_potential}"
                if site.notes:
                    site_text += f"<br/>Notes: {site.notes}"
                elements.append(Paragraph(site_text, styles['Normal']))
                elements.append(Spacer(1, 0.3*cm))
            elements.append(Spacer(1, 0.5*cm))

        # Stands Section
        if forest_plan.stands:
            elements.append(Paragraph("Stands", heading_style))
            stands_data = [['Species', 'Age (years)', 'Size (ha)', 'Development Class']]
            for stand in forest_plan.stands:
                stands_data.append([
                    ', '.join(stand.species),
                    str(stand.age) if stand.age else 'N/A',
                    f"{stand.size_ha:.2f}" if stand.size_ha else 'N/A',
                    stand.development_class or 'N/A'
                ])

            stands_table = Table(stands_data, colWidths=[5*cm, 3*cm, 3*cm, 4*cm])
            stands_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            elements.append(stands_table)
            elements.append(Spacer(1, 1*cm))

        # Management Activities Section
        if forest_plan.management_activities:
            elements.append(Paragraph("Recommended Management Activities", heading_style))
            for i, activity in enumerate(forest_plan.management_activities, 1):
                activity_text = f"<b>{i}. {activity.description}</b><br/>"
                if activity.start_date:
                    activity_text += f"Start: {activity.start_date}"
                if activity.end_date:
                    activity_text += f" - End: {activity.end_date}"
                if activity.responsible_party:
                    activity_text += f"<br/>Responsible: {activity.responsible_party}"
                if activity.notes:
                    activity_text += f"<br/>Notes: {activity.notes}"
                elements.append(Paragraph(activity_text, styles['Normal']))
                elements.append(Spacer(1, 0.3*cm))
            elements.append(Spacer(1, 0.5*cm))

        # Economic Information
        if forest_plan.economic_info:
            elements.append(Paragraph("Economic Information", heading_style))
            econ_data = [
                ['Estimated Value', f"€{forest_plan.economic_info.estimated_value_eur:,.2f}"],
                ['Timber Sales Potential', f"{forest_plan.economic_info.timber_sales_potential_m3:.2f} m³"]
            ]
            if forest_plan.economic_info.notes:
                econ_data.append(['Notes', forest_plan.economic_info.notes])

            econ_table = Table(econ_data, colWidths=[6*cm, 9*cm])
            econ_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f7fafc')),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#2d3748')),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
            ]))
            elements.append(econ_table)

        # Build PDF
        doc.build(elements)
        logger.info(f"PDF report generated: {output_path}")
        return output_path

    def generate_excel_report(self, forest_plan, aggregated_stats: Dict[str, Any],
                             video_path: str, base_name: str) -> Path:
        """Generate Excel report with multiple sheets (requires openpyxl).

        Args:
            forest_plan: ForestPlan object
            aggregated_stats: Aggregated statistics dictionary
            video_path: Path to analyzed video
            base_name: Base filename for output

        Returns:
            Path to generated Excel file
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        output_path = self.output_dir / f"{base_name}.xlsx"

        wb = Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Sheet 1: Summary
        ws_summary = wb.create_sheet("Summary")
        ws_summary.append(['Forest Analysis Report'])
        ws_summary.append(['Plan Name:', forest_plan.name])
        ws_summary.append(['Analysis Date:', aggregated_stats.get('analysis_date', 'N/A')])
        ws_summary.append(['Video Source:', Path(video_path).name])
        ws_summary.append([''])
        ws_summary.append(['Aggregated Statistics'])
        ws_summary.append(['Metric', 'Value'])

        for key, value in aggregated_stats.items():
            if key not in ['species_distribution', 'video_path', 'analysis_date']:
                ws_summary.append([key.replace('_', ' ').title(), value])

        # Format summary sheet
        ws_summary['A1'].font = Font(size=16, bold=True)
        ws_summary['A6'].font = Font(size=14, bold=True)

        # Sheet 2: Sites
        if forest_plan.sites:
            ws_sites = wb.create_sheet("Sites")
            ws_sites.append(['Name', 'Type', 'Nutrient Level', 'Timber Potential', 'Notes'])
            for site in forest_plan.sites:
                ws_sites.append([
                    site.name or '',
                    site.type,
                    site.nutrient_level,
                    site.timber_potential,
                    site.notes or ''
                ])

            # Format header
            for cell in ws_sites[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")

        # Sheet 3: Stands
        if forest_plan.stands:
            ws_stands = wb.create_sheet("Stands")
            ws_stands.append(['Species', 'Age (years)', 'Size (ha)', 'Development Class', 'Notes'])
            for stand in forest_plan.stands:
                ws_stands.append([
                    ', '.join(stand.species),
                    stand.age or '',
                    stand.size_ha or '',
                    stand.development_class or '',
                    stand.notes or ''
                ])

            # Format header
            for cell in ws_stands[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")

        # Sheet 4: Management Activities
        if forest_plan.management_activities:
            ws_activities = wb.create_sheet("Management Activities")
            ws_activities.append(['Description', 'Start Date', 'End Date', 'Responsible Party', 'Notes'])
            for activity in forest_plan.management_activities:
                ws_activities.append([
                    activity.description,
                    str(activity.start_date) if activity.start_date else '',
                    str(activity.end_date) if activity.end_date else '',
                    activity.responsible_party or '',
                    activity.notes or ''
                ])

            # Format header
            for cell in ws_activities[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")

        # Sheet 5: Economic Info
        if forest_plan.economic_info:
            ws_econ = wb.create_sheet("Economic Information")
            ws_econ.append(['Metric', 'Value'])
            ws_econ.append(['Estimated Value (EUR)', forest_plan.economic_info.estimated_value_eur])
            ws_econ.append(['Timber Sales Potential (m³)', forest_plan.economic_info.timber_sales_potential_m3])
            if forest_plan.economic_info.notes:
                ws_econ.append(['Notes', forest_plan.economic_info.notes])

            # Format header
            for cell in ws_econ[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")

        # Save workbook
        wb.save(output_path)
        logger.info(f"Excel report generated: {output_path}")
        return output_path

    def _build_html_content(self, forest_plan, aggregated_stats: Dict[str, Any],
                           video_path: str) -> str:
        """Build HTML content for the report.

        Args:
            forest_plan: ForestPlan object
            aggregated_stats: Aggregated statistics dictionary
            video_path: Path to analyzed video

        Returns:
            HTML content as string
        """
        # Build species distribution HTML
        species_html = ""
        if 'species_distribution' in aggregated_stats and aggregated_stats['species_distribution']:
            species_html = "<h3>Species Distribution</h3><ul>"
            for species, count in aggregated_stats['species_distribution'].items():
                species_html += f"<li><strong>{species}:</strong> {count} trees</li>"
            species_html += "</ul>"

        # Build sites HTML
        sites_html = ""
        if forest_plan.sites:
            sites_html = "<h2>Sites</h2>"
            for i, site in enumerate(forest_plan.sites, 1):
                sites_html += f"""
                <div class="card">
                    <h3>Site {i}: {site.name or 'Unnamed'}</h3>
                    <p><strong>Type:</strong> {site.type}</p>
                    <p><strong>Nutrient Level:</strong> {site.nutrient_level}</p>
                    <p><strong>Timber Potential:</strong> {site.timber_potential}</p>
                    {f'<p><strong>Notes:</strong> {site.notes}</p>' if site.notes else ''}
                </div>
                """

        # Build stands HTML
        stands_html = ""
        if forest_plan.stands:
            stands_html = "<h2>Stands</h2><table><thead><tr>"
            stands_html += "<th>Species</th><th>Age (years)</th><th>Size (ha)</th><th>Development Class</th></tr></thead><tbody>"
            for stand in forest_plan.stands:
                stands_html += f"""
                <tr>
                    <td>{', '.join(stand.species)}</td>
                    <td>{stand.age if stand.age else 'N/A'}</td>
                    <td>{f'{stand.size_ha:.2f}' if stand.size_ha else 'N/A'}</td>
                    <td>{stand.development_class or 'N/A'}</td>
                </tr>
                """
            stands_html += "</tbody></table>"

        # Build activities HTML
        activities_html = ""
        if forest_plan.management_activities:
            activities_html = "<h2>Recommended Management Activities</h2>"
            for i, activity in enumerate(forest_plan.management_activities, 1):
                activities_html += f"""
                <div class="card">
                    <h3>{i}. {activity.description}</h3>
                    {f'<p><strong>Start Date:</strong> {activity.start_date}</p>' if activity.start_date else ''}
                    {f'<p><strong>End Date:</strong> {activity.end_date}</p>' if activity.end_date else ''}
                    {f'<p><strong>Responsible:</strong> {activity.responsible_party}</p>' if activity.responsible_party else ''}
                    {f'<p><strong>Notes:</strong> {activity.notes}</p>' if activity.notes else ''}
                </div>
                """

        # Build economic info HTML
        econ_html = ""
        if forest_plan.economic_info:
            econ_html = f"""
            <h2>Economic Information</h2>
            <div class="card">
                <p><strong>Estimated Value:</strong> €{forest_plan.economic_info.estimated_value_eur:,.2f}</p>
                <p><strong>Timber Sales Potential:</strong> {forest_plan.economic_info.timber_sales_potential_m3:.2f} m³</p>
                {f'<p><strong>Notes:</strong> {forest_plan.economic_info.notes}</p>' if forest_plan.economic_info.notes else ''}
            </div>
            """

        # Complete HTML template
        html = f"""
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Forest Analysis Report - {forest_plan.name}</title>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            padding: 20px;
            color: #333;
        }}
        .container {{
            max-width: 1200px;
            margin: 0 auto;
            background: white;
            border-radius: 16px;
            box-shadow: 0 20px 60px rgba(0, 0, 0, 0.3);
            overflow: hidden;
        }}
        .header {{
            background: linear-gradient(135deg, #2d3748 0%, #1a202c 100%);
            color: white;
            padding: 40px;
            text-align: center;
        }}
        .header h1 {{ font-size: 32px; margin-bottom: 10px; }}
        .header p {{ font-size: 14px; color: #cbd5e0; }}
        .content {{ padding: 40px; }}
        h2 {{
            color: #2d3748;
            font-size: 24px;
            margin: 30px 0 20px 0;
            padding-bottom: 10px;
            border-bottom: 3px solid #667eea;
        }}
        h3 {{ color: #2d3748; font-size: 18px; margin: 15px 0 10px 0; }}
        .stats-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(250px, 1fr));
            gap: 20px;
            margin: 20px 0;
        }}
        .stat-card {{
            background: #f7fafc;
            padding: 20px;
            border-radius: 12px;
            border-left: 4px solid #667eea;
        }}
        .stat-card h4 {{
            color: #718096;
            font-size: 12px;
            text-transform: uppercase;
            margin-bottom: 8px;
        }}
        .stat-card p {{
            color: #2d3748;
            font-size: 24px;
            font-weight: bold;
        }}
        .card {{
            background: #f7fafc;
            padding: 20px;
            border-radius: 12px;
            margin: 15px 0;
            border: 1px solid #e2e8f0;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            margin: 20px 0;
            background: white;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }}
        th, td {{
            padding: 12px;
            text-align: left;
            border-bottom: 1px solid #e2e8f0;
        }}
        th {{
            background: #667eea;
            color: white;
            font-weight: 600;
        }}
        tr:hover {{ background: #f7fafc; }}
        ul {{ margin: 15px 0; padding-left: 20px; }}
        li {{ margin: 8px 0; }}
        @media print {{
            body {{ background: white; padding: 0; }}
            .container {{ box-shadow: none; }}
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>🌲 Forest Analysis Report</h1>
            <p>{forest_plan.name}</p>
            <p>Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
        </div>

        <div class="content">
            <h2>Executive Summary</h2>
            <div class="stats-grid">
                <div class="stat-card">
                    <h4>Total Trees Detected</h4>
                    <p>{aggregated_stats.get('total_trees_detected', 0)}</p>
                </div>
                <div class="stat-card">
                    <h4>Avg Trees per Frame</h4>
                    <p>{aggregated_stats.get('avg_trees_per_frame', 0):.2f}</p>
                </div>
                <div class="stat-card">
                    <h4>Canopy Coverage</h4>
                    <p>{aggregated_stats.get('avg_canopy_coverage_percent', 0):.2f}%</p>
                </div>
                <div class="stat-card">
                    <h4>Vegetation Density</h4>
                    <p>{aggregated_stats.get('avg_vegetation_density_percent', 0):.2f}%</p>
                </div>
            </div>

            {species_html}
            {sites_html}
            {stands_html}
            {activities_html}
            {econ_html}

            <h2>Analysis Metadata</h2>
            <div class="card">
                <p><strong>Video Source:</strong> {Path(video_path).name}</p>
                <p><strong>Frames Analyzed:</strong> {aggregated_stats.get('total_frames_analyzed', 0)}</p>
                <p><strong>Analysis Date:</strong> {aggregated_stats.get('analysis_date', 'N/A')}</p>
            </div>
        </div>
    </div>
</body>
</html>
        """

        return html

