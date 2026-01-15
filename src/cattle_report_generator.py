"""
Cattle Report Generator
=======================

Generate professional reports for cattle monitoring and herd management.

Supports multiple formats:
- HTML: Interactive web-based reports
- CSV: Data exports for analysis
- PDF: Professional documents (requires reportlab)
- Excel: Multi-sheet workbooks (requires openpyxl)

Author: Adrian
Date: December 2024
"""

from pathlib import Path
from typing import Dict, Any, Optional, Union
from datetime import datetime
import logging

from .cattle_structures import CattlePlan, HerdStatistics

logger = logging.getLogger(__name__)


class CattleReportGenerator:
    """Generate professional cattle monitoring reports."""
    
    def __init__(self, output_dir: str = "output"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        logger.info(f"CattleReportGenerator initialized with output_dir: {output_dir}")
    
    def generate_all_reports(self, cattle_plan: CattlePlan,
                           aggregated_stats: Dict[str, Any],
                           video_path: Union[str, int]) -> Dict[str, Path]:
        """Generate all available report formats."""
        # Convert video_path to string for display
        video_path_str = str(video_path)

        report_paths = {}

        # Always generate HTML and CSV (no dependencies)
        try:
            report_paths['html'] = self.generate_html_report(cattle_plan, aggregated_stats, video_path_str)
        except Exception as e:
            logger.error(f"HTML report generation failed: {e}")

        try:
            csv_paths = self.generate_csv_reports(cattle_plan, aggregated_stats)
            report_paths.update(csv_paths)
        except Exception as e:
            logger.error(f"CSV report generation failed: {e}")

        # Try PDF (requires reportlab)
        try:
            report_paths['pdf'] = self.generate_pdf_report(cattle_plan, aggregated_stats, video_path_str)
        except ImportError:
            logger.warning("PDF generation skipped: reportlab not installed")
        except Exception as e:
            logger.error(f"PDF report generation failed: {e}")

        # Try Excel (requires openpyxl)
        try:
            report_paths['excel'] = self.generate_excel_report(cattle_plan, aggregated_stats, video_path_str)
        except ImportError:
            logger.warning("Excel generation skipped: openpyxl not installed")
        except Exception as e:
            logger.error(f"Excel report generation failed: {e}")

        logger.info(f"Generated {len(report_paths)} report formats")
        return report_paths
    
    def generate_html_report(self, cattle_plan: CattlePlan,
                            aggregated_stats: Dict[str, Any],
                            video_path: str) -> Path:
        """Generate interactive HTML report."""
        output_path = self.output_dir / "cattle_report.html"
        
        stats = cattle_plan.herd_stats
        
        html_content = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Cattle Monitoring Report</title>
    <style>
        body {{ font-family: Arial, sans-serif; margin: 40px; background: #f5f5f5; }}
        .container {{ max-width: 1200px; margin: 0 auto; background: white; padding: 40px; border-radius: 8px; box-shadow: 0 2px 8px rgba(0,0,0,0.1); }}
        h1 {{ color: #2c3e50; border-bottom: 3px solid #3498db; padding-bottom: 10px; }}
        h2 {{ color: #34495e; margin-top: 30px; }}
        .stats-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 20px; margin: 20px 0; }}
        .stat-card {{ background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 20px; border-radius: 8px; }}
        .stat-value {{ font-size: 32px; font-weight: bold; }}
        .stat-label {{ font-size: 14px; opacity: 0.9; margin-top: 5px; }}
        table {{ width: 100%; border-collapse: collapse; margin: 20px 0; }}
        th, td {{ padding: 12px; text-align: left; border-bottom: 1px solid #ddd; }}
        th {{ background: #34495e; color: white; }}
        .recommendation {{ background: #e8f5e9; padding: 15px; border-left: 4px solid #4caf50; margin: 10px 0; }}
        .alert {{ background: #fff3cd; padding: 15px; border-left: 4px solid #ffc107; margin: 10px 0; }}
    </style>
</head>
<body>
    <div class="container">
        <h1>🐄 Cattle Monitoring Report</h1>
        <p><strong>Generated:</strong> {cattle_plan.generated_at}</p>
        <p><strong>Video Source:</strong> {video_path}</p>
        <p><strong>Analysis Mode:</strong> {cattle_plan.analysis_mode}</p>
        
        <h2>Herd Statistics</h2>
        <div class="stats-grid">
            <div class="stat-card">
                <div class="stat-value">{stats.total_cattle_detected if stats else 0}</div>
                <div class="stat-label">Total Cattle Detected</div>
            </div>
            <div class="stat-card">
                <div class="stat-value">{stats.unique_cattle_tracked if stats else 0}</div>
                <div class="stat-label">Unique Cattle Tracked</div>
            </div>
            <div class="stat-card">
                <div class="stat-value">{stats.avg_cattle_per_frame:.1f if stats else 0}</div>
                <div class="stat-label">Avg Cattle per Frame</div>
            </div>
            <div class="stat-card">
                <div class="stat-value">{stats.max_cattle_in_frame if stats else 0}</div>
                <div class="stat-label">Max in Single Frame</div>
            </div>
        </div>
        
        <h2>Behavior Distribution</h2>
        <table>
            <tr><th>Behavior</th><th>Percentage</th></tr>
"""
        
        if stats and stats.behavior_distribution:
            for behavior, percentage in stats.behavior_distribution.items():
                html_content += f"            <tr><td>{behavior.capitalize()}</td><td>{percentage:.1f}%</td></tr>\n"
        
        html_content += f"""        </table>
        
        <h2>Health Status</h2>
        <div class="stats-grid">
            <div class="stat-card" style="background: linear-gradient(135deg, #48bb78 0%, #38a169 100%);">
                <div class="stat-value">{stats.healthy_percentage:.1f}% if stats else 0}%</div>
                <div class="stat-label">Healthy</div>
            </div>
            <div class="stat-card" style="background: linear-gradient(135deg, #ed8936 0%, #dd6b20 100%);">
                <div class="stat-value">{stats.attention_needed_percentage:.1f if stats else 0}%</div>
                <div class="stat-label">Attention Needed</div>
            </div>
        </div>
        
        <h2>Management Recommendations</h2>
"""
        
        for rec in cattle_plan.management_recommendations:
            html_content += f'        <div class="recommendation">{rec}</div>\n'
        
        if cattle_plan.health_alerts:
            html_content += "        <h2>Health Alerts</h2>\n"
            for alert in cattle_plan.health_alerts:
                html_content += f'        <div class="alert">{alert}</div>\n'
        
        html_content += """    </div>
</body>
</html>"""
        
        output_path.write_text(html_content, encoding='utf-8')
        logger.info(f"HTML report saved to: {output_path}")
        return output_path

    def generate_csv_reports(self, cattle_plan: CattlePlan,
                            aggregated_stats: Dict[str, Any]) -> Dict[str, Path]:
        """Generate CSV data exports."""
        import csv

        csv_paths = {}

        # Statistics CSV
        stats_path = self.output_dir / "cattle_statistics.csv"
        with open(stats_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(['Metric', 'Value'])

            stats = cattle_plan.herd_stats
            if stats:
                writer.writerow(['Total Cattle Detected', stats.total_cattle_detected])
                writer.writerow(['Unique Cattle Tracked', stats.unique_cattle_tracked])
                writer.writerow(['Avg Cattle per Frame', f'{stats.avg_cattle_per_frame:.2f}'])
                writer.writerow(['Max Cattle in Frame', stats.max_cattle_in_frame])
                writer.writerow(['Min Cattle in Frame', stats.min_cattle_in_frame])
                writer.writerow(['Healthy Percentage', f'{stats.healthy_percentage:.1f}%'])
                writer.writerow(['Attention Needed Percentage', f'{stats.attention_needed_percentage:.1f}%'])
                writer.writerow(['Active Percentage', f'{stats.active_percentage:.1f}%'])
                writer.writerow(['Resting Percentage', f'{stats.resting_percentage:.1f}%'])

        csv_paths['statistics_csv'] = stats_path
        logger.info(f"Statistics CSV saved to: {stats_path}")

        # Behavior distribution CSV
        behavior_path = self.output_dir / "cattle_behavior.csv"
        with open(behavior_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(['Behavior', 'Percentage'])

            if cattle_plan.herd_stats and cattle_plan.herd_stats.behavior_distribution:
                for behavior, percentage in cattle_plan.herd_stats.behavior_distribution.items():
                    writer.writerow([behavior.capitalize(), f'{percentage:.1f}%'])

        csv_paths['behavior_csv'] = behavior_path
        logger.info(f"Behavior CSV saved to: {behavior_path}")

        # Tracks CSV
        tracks_path = self.output_dir / "cattle_tracks.csv"
        with open(tracks_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(['Track ID', 'First Frame', 'Last Frame', 'Duration (frames)',
                           'Dominant Behavior', 'Health Status'])

            for track in cattle_plan.cattle_tracks:
                duration = track.last_seen_frame - track.first_seen_frame + 1
                writer.writerow([
                    track.track_id,
                    track.first_seen_frame,
                    track.last_seen_frame,
                    duration,
                    track.dominant_behavior.value,
                    track.health_status.value
                ])

        csv_paths['tracks_csv'] = tracks_path
        logger.info(f"Tracks CSV saved to: {tracks_path}")

        return csv_paths

    def generate_pdf_report(self, cattle_plan: CattlePlan,
                           aggregated_stats: Dict[str, Any],
                           video_path: str) -> Path:
        """Generate PDF report (requires reportlab)."""
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

        output_path = self.output_dir / "cattle_report.pdf"
        doc = SimpleDocTemplate(str(output_path), pagesize=A4)
        story = []
        styles = getSampleStyleSheet()

        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#2c3e50'),
            spaceAfter=30
        )
        story.append(Paragraph("🐄 Cattle Monitoring Report", title_style))
        story.append(Spacer(1, 0.2*inch))

        # Metadata
        story.append(Paragraph(f"<b>Generated:</b> {cattle_plan.generated_at}", styles['Normal']))
        story.append(Paragraph(f"<b>Video Source:</b> {video_path}", styles['Normal']))
        story.append(Paragraph(f"<b>Analysis Mode:</b> {cattle_plan.analysis_mode}", styles['Normal']))
        story.append(Spacer(1, 0.3*inch))

        # Herd Statistics
        story.append(Paragraph("Herd Statistics", styles['Heading2']))
        stats = cattle_plan.herd_stats
        if stats:
            stats_data = [
                ['Metric', 'Value'],
                ['Total Cattle Detected', str(stats.total_cattle_detected)],
                ['Unique Cattle Tracked', str(stats.unique_cattle_tracked)],
                ['Avg Cattle per Frame', f'{stats.avg_cattle_per_frame:.2f}'],
                ['Max Cattle in Frame', str(stats.max_cattle_in_frame)],
                ['Healthy Percentage', f'{stats.healthy_percentage:.1f}%'],
                ['Attention Needed', f'{stats.attention_needed_percentage:.1f}%']
            ]

            stats_table = Table(stats_data, colWidths=[3*inch, 2*inch])
            stats_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#34495e')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(stats_table)

        story.append(Spacer(1, 0.3*inch))

        # Recommendations
        story.append(Paragraph("Management Recommendations", styles['Heading2']))
        for rec in cattle_plan.management_recommendations:
            story.append(Paragraph(f"• {rec}", styles['Normal']))
            story.append(Spacer(1, 0.1*inch))

        # Build PDF
        doc.build(story)
        logger.info(f"PDF report saved to: {output_path}")
        return output_path

    def generate_excel_report(self, cattle_plan: CattlePlan,
                             aggregated_stats: Dict[str, Any],
                             video_path: str) -> Path:
        """Generate Excel workbook (requires openpyxl)."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        output_path = self.output_dir / "cattle_report.xlsx"
        wb = Workbook()

        # Summary sheet
        ws = wb.active
        ws.title = "Summary"

        ws['A1'] = "Cattle Monitoring Report"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A3'] = "Generated:"
        ws['B3'] = cattle_plan.generated_at
        ws['A4'] = "Video Source:"
        ws['B4'] = video_path

        stats = cattle_plan.herd_stats
        if stats:
            ws['A6'] = "Herd Statistics"
            ws['A6'].font = Font(size=14, bold=True)

            ws['A7'] = "Total Cattle Detected"
            ws['B7'] = stats.total_cattle_detected
            ws['A8'] = "Unique Cattle Tracked"
            ws['B8'] = stats.unique_cattle_tracked
            ws['A9'] = "Avg Cattle per Frame"
            ws['B9'] = round(stats.avg_cattle_per_frame, 2)
            ws['A10'] = "Healthy Percentage"
            ws['B10'] = f"{stats.healthy_percentage:.1f}%"

        # Tracks sheet
        ws_tracks = wb.create_sheet("Tracks")
        ws_tracks['A1'] = "Track ID"
        ws_tracks['B1'] = "First Frame"
        ws_tracks['C1'] = "Last Frame"
        ws_tracks['D1'] = "Dominant Behavior"
        ws_tracks['E1'] = "Health Status"

        for i, track in enumerate(cattle_plan.cattle_tracks, start=2):
            ws_tracks[f'A{i}'] = track.track_id
            ws_tracks[f'B{i}'] = track.first_seen_frame
            ws_tracks[f'C{i}'] = track.last_seen_frame
            ws_tracks[f'D{i}'] = track.dominant_behavior.value
            ws_tracks[f'E{i}'] = track.health_status.value

        wb.save(output_path)
        logger.info(f"Excel report saved to: {output_path}")
        return output_path


